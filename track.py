import cv2
from ultralytics import YOLO
import pandas as pd
from collections import defaultdict, Counter
import numpy as np
import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Путь к модели и видео
model_path = 'best.pt'
video_path = 'test.mp4'

# Список классов
classes = ['Euro6_Class1', 'Euro6_Class2', 'Euro6_Class3', 'Euro6_Class4', 
           'Euro6_Class5', 'Euro6_Class6', 'Other']

# Загрузка модели
model = YOLO(model_path)

# Видео
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
    print("Ошибка открытия видеофайла")
    exit()

# Параметры
object_tracker = {}
object_id_counter = 0
class_counter = Counter()
trajectory_length = 5
direction_threshold = -20

prev_time = time.time()

while True:
    ret, frame = cap.read()
    if not ret:
        break

    results = model.predict(frame, conf=0.4, iou=0.5, verbose=False)
    detections = results[0].boxes.xyxy.cpu().numpy() if results[0].boxes.xyxy is not None else []
    confs = results[0].boxes.conf.cpu().numpy() if results[0].boxes.conf is not None else []
    cls_ids = results[0].boxes.cls.cpu().numpy() if results[0].boxes.cls is not None else []

    current_objects = []

    for i, det in enumerate(detections):
        x1, y1, x2, y2 = det
        conf = confs[i]
        cls_id = int(cls_ids[i])
        label = classes[cls_id]
        center = ((x1 + x2) // 2, (y1 + y2) // 2)

        matched_id = None
        for obj_id, data in object_tracker.items():
            prev_center = data['trajectory'][-1]
            if np.linalg.norm(np.array(center) - np.array(prev_center)) < 50:
                matched_id = obj_id
                break

        if matched_id is None:
            object_id_counter += 1
            matched_id = object_id_counter
            object_tracker[matched_id] = {
                'trajectory': [center],
                'labels': [label],
                'counted': False
            }
        else:
            object_tracker[matched_id]['trajectory'].append(center)
            object_tracker[matched_id]['labels'].append(label)
            if len(object_tracker[matched_id]['trajectory']) > trajectory_length:
                object_tracker[matched_id]['trajectory'].pop(0)
                object_tracker[matched_id]['labels'].pop(0)

        current_objects.append(matched_id)

        color = (0, 255, 0)
        cv2.rectangle(frame, (int(x1), int(y1)), (int(x2), int(y2)), color, 2)
        cv2.putText(frame, f'ID {matched_id} {label}', (int(x1), int(y1) - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

    for obj_id, data in list(object_tracker.items()):
        if obj_id not in current_objects:
            continue

        if len(data['trajectory']) >= trajectory_length and not data['counted']:
            dy = data['trajectory'][0][1] - data['trajectory'][-1][1]
            if dy < direction_threshold:
                most_common_label = Counter(data['labels']).most_common(1)[0][0]
                class_counter[most_common_label] += 1
                data['counted'] = True

    to_delete = [obj_id for obj_id in object_tracker if obj_id not in current_objects]
    for obj_id in to_delete:
        del object_tracker[obj_id]

    y_offset = 20
    for cls, count in class_counter.items():
        cv2.putText(frame, f'{cls}: {count}', (10, y_offset),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        y_offset += 25

    curr_time = time.time()
    fps = 1 / (curr_time - prev_time)
    prev_time = curr_time
    cv2.putText(frame, f'FPS: {fps:.2f}', (frame.shape[1] - 120, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 0), 2)

    cv2.imshow('Detection', frame)
    key = cv2.waitKey(1) & 0xFF
    if key == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

# Формируем таблицу по шаблону
data = []
total = 0
for i in range(6):
    count = class_counter.get(f'Euro6_Class{i+1}', 0)
    data.append(count)
    total += count

other_count = class_counter.get('Other', 0)
total += other_count

df = pd.DataFrame({
    'Класс': ['Класс 1', 'Класс 2', 'Класс 3', 'Класс 4', 'Класс 5', 'Класс 6', 'Остальные', 'Итого'],
    'Количество': data + [other_count, total]
})

# Сохраняем красиво в Excel
excel_path = 'report.xlsx'
df.to_excel(excel_path, index=False, startrow=4)

# Украшаем заголовок
wb = load_workbook(excel_path)
ws = wb.active

# Добавляем заголовок
ws.merge_cells('A1:B1')
ws['A1'] = 'Отчет о интенсивности движения транспорта через пункт'
ws['A1'].font = Font(size=14, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

wb.save(excel_path)
print(f'Отчёт сохранён в {excel_path}')
